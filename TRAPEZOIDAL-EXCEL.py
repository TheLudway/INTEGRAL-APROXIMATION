import matplotlib.pyplot as plt
from openpyxl import  Workbook, load_workbook
from openpyxl.utils import get_column_letter


def simpson(a = 3, b = 40982, wb = load_workbook('SPEED.xlsx')):
    ws = wb.active
    Resultado_suma = 0
    Delta_x = 0
    Resultado_final = 0
    for row in range(a, b):
        for col in [2]:
            char = get_column_letter(col)
            Resultado_suma += (ws[char + str(row)].value) * 2

    Resultado_suma += ws["B2"].value + ws["B40981"].value
    Delta_x = (ws["A40981"].value - ws["A2"].value) / 40980
    Resultado_final = (Resultado_suma * Delta_x) / 2
    print(Resultado_final)


def graficar_S(x = [], y = [], wb = load_workbook('SPEED.xlsx'), a = 3, b = 40982):
    ws = wb.active
    for row in range(a, b):
        for col in [2]:
            char = get_column_letter(col)
            y.append(ws[char + str(row)].value)
        for col in [1]:
            char = get_column_letter(col)
            x.append(ws[char + str(row)].value)
    plt.plot(x, y, color = (32/255, 32/255, 32/255))
    plt.title("Velocidad del automóvil a lo largo de todo el viaje")
    plt.xlabel("Tiempo t(s)")
    plt.ylabel("Velocidad v(m/s)")
    plt.fill_between(x, y, color = (192/255, 192/255, 192/255))
    plt.show()

simpson()
graficar_S()


"""
⠀⠀⠀⠀⠀⠀⠀⠀⠀⢠⣿⣶⣄⣀⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⢀⣴⣿⣿⣿⣿⣿⣿⣿⣿⣿⣶⣦⣄⣀⡀⣠⣾⡇⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⣴⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀⠀⠀
⠀⠀⠀⠀⢀⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠿⢿⣿⣿⡇⠀⠀⠀⠀
⠀⣶⣿⣦⣜⣿⣿⣿⡟⠻⣿⣿⣿⣿⣿⣿⣿⡿⢿⡏⣴⣺⣦⣙⣿⣷⣄⠀⠀⠀
⠀⣯⡇⣻⣿⣿⣿⣿⣷⣾⣿⣬⣥⣭⣽⣿⣿⣧⣼⡇⣯⣇⣹⣿⣿⣿⣿⣧⠀⠀
⠀⠹⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠸⣿⣿⣿⣿⣿⣿⣿⣷⠀
""" 
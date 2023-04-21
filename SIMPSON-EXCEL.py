import matplotlib.pyplot as plt
from openpyxl import  Workbook, load_workbook
from openpyxl.utils import get_column_letter


def simpson(a = 3, b = 40982, wb = load_workbook('SPEED.xlsx')):
    ws = wb.active
    Resultado_suma = 0
    Delta_x = 0
    Resultado_final = 0
    for row in range(a, b):
        if row % 2 == 1:
            for col in [2]:
                char = get_column_letter(col)
                Resultado_suma += (ws[char + str(row)].value) * 4
        else:
            for col in [2]:
                char = get_column_letter(col)
                Resultado_suma += (ws[char + str(row)].value) * 2
    Resultado_suma += ws["B2"].value + ws["B40981"].value
    Delta_x = (ws["A40981"].value - ws["A2"].value) / 40980
    Resultado_final = (Resultado_suma * Delta_x) / 3
    print(Resultado_final)
    Redondear = round(Resultado_final, 2)
    return Redondear


def graficar_S(x = [], y = [], wb = load_workbook('SPEED.xlsx'), a = 3, b = 40982):
    ws = wb.active
    for row in range(a, b):
        for col in [2]:
            char = get_column_letter(col)
            y.append(ws[char + str(row)].value)
        for col in [1]:
            char = get_column_letter(col)
            x.append(ws[char + str(row)].value)
    plt.plot(x, y, color = "black")
    plt.title(f"Velocidad del automóvil a lo largo de todo el viaje {simpson()}")
    plt.xlabel("Tiempo t(s)")
    plt.ylabel("Velocidad v(m/s)")
    plt.fill_between(x, y, color = (232/255, 238/255, 249/255))
    plt.show()


graficar_S()



"""
⠀⠀⠀⠀⠀⠀⠀⠀⠀⢠⣿⣶⣄⣀⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⢀⣴⣿⣿⣿⣿⣿⣿⣿⣿⣿⣶⣦⣄⣀⡀⣠⣾⡇⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⣴⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀⠀⠀
⠀⠀⠀⠀⢀⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠿⢿⣿⣿⡇⠀⠀⠀⠀
⠀⣶⣿⣦⣜⣿⣿⣿⡟⠻⣿⣿⣿⣿⣿⣿⣿⡿⢿⡏⣴⣺⣦⣙⣿⣷⣄⠀⠀⠀
⠀⣯⡇⣻⣿⣿⣿⣿⣷⣾⣿⣬⣥⣭⣽⣿⣿⣧⣼⡇⣯⣇⣹⣿⣿⣿⣿⣧⠀⠀
⠀⠹⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠸⣿⣿⣿⣿⣿⣿⣿⣷⠀
"""
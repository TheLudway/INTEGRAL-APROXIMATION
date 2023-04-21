import matplotlib.pyplot as plt
from openpyxl import  Workbook, load_workbook
from openpyxl.utils import get_column_letter


def Riemann_R(a = 3, b = 40982, wb = load_workbook('SPEED.xlsx')):
    ws = wb.active
    Resultado_suma = 0
    Delta_x = 0
    Resultado_final = 0
    for row in range(a, b):
        for col in [2]:
            char = get_column_letter(col)
            Resultado_suma += (ws[char + str(row)].value)

    Delta_x = (ws["A40981"].value - ws["A2"].value) / 40980
    Resultado_final = (Resultado_suma * Delta_x) 
    print(Resultado_final)
    return Resultado_final


def graficar_RR(x = [], y = [], wb = load_workbook('SPEED.xlsx'), a = 3, b = 40982):
    ws = wb.active
    for row in range(a, b):
        for col in [2]:
            char = get_column_letter(col)
            y.append(ws[char + str(row)].value)
        for col in [1]:
            char = get_column_letter(col)
            x.append(ws[char + str(row)].value)
    plt.plot(x, y, color = (32/255, 32/255, 32/255))
    plt.title("Velocidad del automóvil a lo largo de todo el viaje")
    plt.xlabel("Tiempo t(s)")
    plt.ylabel("Velocidad v(m/s)")
    plt.fill_between(x, y, color = (192/255, 192/255, 192/255))
    plt.show()

Riemann_R()
graficar_RR()


"""
⠀⠀⠀⠀⠀⠀⠀⠀⠀⢠⣿⣶⣄⣀⡀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⠀⢀⣴⣿⣿⣿⣿⣿⣿⣿⣿⣿⣶⣦⣄⣀⡀⣠⣾⡇⠀⠀⠀⠀
⠀⠀⠀⠀⠀⠀⣴⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⠀⠀⠀
⠀⠀⠀⠀⢀⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠿⢿⣿⣿⡇⠀⠀⠀⠀
⠀⣶⣿⣦⣜⣿⣿⣿⡟⠻⣿⣿⣿⣿⣿⣿⣿⡿⢿⡏⣴⣺⣦⣙⣿⣷⣄⠀⠀⠀
⠀⣯⡇⣻⣿⣿⣿⣿⣷⣾⣿⣬⣥⣭⣽⣿⣿⣧⣼⡇⣯⣇⣹⣿⣿⣿⣿⣧⠀⠀
⠀⠹⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠸⣿⣿⣿⣿⣿⣿⣿⣷⠀
""" 